'''
Created on May 8, 2020

@author: Home
'''
from Excel_utils2 import Excel_utils2
from openpyxl.styles import Font

if __name__ == '__main__':
    data_in = Excel_utils2('C:/temp/test1.xlsx', 'symbols', 'in')
    
    print (data_in.sht.max_row)
    
    i=0
    for row in data_in.sht.rows:
        i+=1
        if (i == 1): # skip the 1st row, it contains headings
            continue
        print (data_in.sht.cell(row=i, column =1).value)
        
    for value in data_in.sht.iter_rows(values_only=True):
       print (value)
       
    print("Data In:")    
    print(data_in.__dict__)
    print('Workbook: ', data_in.wkbk)
    print('Sheet: ', data_in.sht)
    
    print("Data Out:")   
    data_out = Excel_utils2(' ', 'Quotes', 'out')
    print(data_out.__dict__)
    print('Workbook: ', data_out.wkbk)
    print('Sheet: ', data_out.sht)
    i, j = 1,1
    data_out.set_cell(i, j, 'ABC123', "Arial", True, 12)
    i,j  = 2,2
    data_out.set_cell(i, j, 'DEF456', "Courier", False, 10)
    data_out.wkbk.save('C:/temp/abc123.xlsx')