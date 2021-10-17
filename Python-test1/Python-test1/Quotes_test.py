'''
Created on 03Apr2020

@author: Home
'''
import openpyxl
from openpyxl.styles import NamedStyle, Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from Tools.scripts.gprof2html import header

wb_in = openpyxl.load_workbook('C:/temp/test-parameters.xlsx')
wksh_in = wb_in['test_parameters'] # Set the work sheet to the test parameters
# Get account # and password
target_url = wksh_in.cell( 2,  2).value
acc_num = wksh_in.cell( 3, 2).value
passwd = wksh_in.cell( 4, 2).value

browser = "C:\\Selenium\\chromedriver.exe"
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--incognito")

# Open Chrome and navigate to the test website
driver = webdriver.Chrome(browser)
driver.get(target_url)
page_title = "BMO InvestorLine" 
assert page_title in driver.title
print ("Loading the target page...")

# Enter account #
log_in = driver.find_element_by_id("loginText")
log_in.send_keys(acc_num)

# Enter password
password = driver.find_element_by_name("password")
password.send_keys(passwd)

# Click log in button to log in
# driver.implicitly_wait(10) # seconds
log_in_button = driver.find_element_by_id("sasi_btn").click()

wait = WebDriverWait(driver, 10)

wksh_in = wb_in['CA_Symbols'] # Set the work sheet to the test stock symbols
  
wb_out = openpyxl.Workbook() # Create a new work book and work sheet
wksh_out = wb_out.active

header = NamedStyle(name="header")
header.font = Font(name = 'Arial', bold=True, size = 12)

normal = NamedStyle(name="normal")
normal.font = Font(name='Arial', bold = False, size = 10)

wksh_out.cell(1, 1).value = "Symbol"
wksh_out.cell(1, 1).style = header
wksh_out.cell(1, 2).value = "Last"
wksh_out.cell(1, 2).style = header
wksh_out.cell(1, 3).value = "Bid"
wksh_out.cell(1, 3).style = header
wksh_out.cell(1, 4).value = "Ask"
wksh_out.cell(1, 4).style = header
wksh_out.cell(1, 5).value = "High"
wksh_out.cell(1, 5).style = header
wksh_out.cell(1, 6).value = "Low"
wksh_out.cell(1, 6).style = header
wksh_out.cell(1, 7).value = "Open"
wksh_out.cell(1, 7).style = header
wksh_out.cell(1, 8).value = "Pre.Close"
wksh_out.cell(1, 8).style = header

i=0
for row in wksh_in.rows:
     i+=1
     if (i == 1): # skip the 1st row, it contains headings
        continue

     # Wait for the Quotes & Tools menu item to appear
     wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#nav_quotes_main > a:nth-child(1)")))
     driver.find_element_by_css_selector("#nav_quotes_main > a:nth-child(1)").click() # click on Quotes & Tools
     # Wait for the Symbol text box to appear
     wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.paddingLeft25:nth-child(2) > input:nth-child(2)")))
     symbol = driver.find_element_by_css_selector("div.paddingLeft25:nth-child(2) > input:nth-child(2)")
     test_symbol = row[0].value
     symbol.send_keys(test_symbol, Keys.RETURN) # enter a symbol and get a quote
      
     # get a quote but wait for the "Summary" tab to appear
     wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".page_navigation > div:nth-child(1) > a:nth-child(1)")))
        
     last_price = driver.find_element_by_css_selector(".brownContainer > div:nth-child(1) > span:nth-child(1)").get_attribute("innerText")
     bid_price = driver.find_element_by_css_selector("#table1 > table > tbody > tr.tableRowBG > td:nth-child(2)").get_attribute("innerText")
     ask_price = driver.find_element_by_css_selector("#table1 > table > tbody > tr:nth-child(2) > td:nth-child(2)").get_attribute("innerText")
     high = driver.find_element_by_css_selector("#table1 > table > tbody > tr.tableRowBG > td:nth-child(4)").get_attribute("innerText")
     low =  driver.find_element_by_css_selector("#table1 > table > tbody > tr:nth-child(2) > td:nth-child(4)").get_attribute("innerText")
     open_price = driver.find_element_by_css_selector("#table1 > table > tbody > tr.tableRowBG > td:nth-child(6)").get_attribute("innerText")
     prev_close = driver.find_element_by_css_selector("#table1 > table > tbody > tr:nth-child(2) > td:nth-child(6)").get_attribute("innerText")
    
     wksh_out.cell( i, 1).value = test_symbol
     wksh_out.cell( i, 1).style = normal
     wksh_out.cell( i, 2).value = last_price
     wksh_out.cell( i, 2).style = normal
     wksh_out.cell( i, 3).value =bid_price
     wksh_out.cell( i, 3).style = normal
     wksh_out.cell( i, 4).value =ask_price
     wksh_out.cell( i, 4).style = normal
     wksh_out.cell( i, 5).value =high
     wksh_out.cell( i, 5).style = normal
     wksh_out.cell( i, 6).value =low
     wksh_out.cell( i, 6).style = normal
     wksh_out.cell( i, 7).value =open_price
     wksh_out.cell( i, 7).style = normal
     wksh_out.cell( i, 8).value =prev_close
     wksh_out.cell( i, 8).style = normal
    
wksh_out.title = 'Quotes'
wb_out.save("C:/temp/Quotes-test-simple.xlsx")
print ("Test ending ...")
driver.quit()