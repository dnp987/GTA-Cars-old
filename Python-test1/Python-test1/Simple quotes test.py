'''
Created on 03Apr2020

@author: Home
'''
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

wb = openpyxl.load_workbook(r'C:\temp\test1.xlsx')
wksh = wb['test_parameters']
# Get account # and password
target_url = wksh.cell(row = 2, column = 2).value
acc_num = wksh.cell(row = 3,column = 2).value
passwd = wksh.cell(row = 4,column = 2).value

browser = "C:\\Selenium\\chromedriver.exe"
page_title = "BMO InvestorLine" 
chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument("--incognito")

# Open Chrome and navigate to the test website
driver = webdriver.Chrome(browser)
driver.get(target_url)

assert page_title in driver.title
print ("Loading the target page...")

# Enter account #
log_in = driver.find_element_by_id("loginText")
log_in.send_keys(acc_num)

# Enter password
password = driver.find_element_by_name("password")
password.send_keys(passwd)

# Click log in button to log in
# driver.implicitly_wait(10) # seconds
log_in_button = driver.find_element_by_id("sasi_btn").click()

wait = WebDriverWait(driver, 10)

wksh = wb['symbols'] 
wksh.cell(row=1, column=1).value = "Symbol"
wksh.cell(row=1, column=2).value = "Last"
wksh.cell(row=1, column=3).value = "Bid"
wksh.cell(row=1, column=4).value = "Ask"
wksh.cell(row=1, column=5).value = "High"
wksh.cell(row=1, column=6).value = "Low"
wksh.cell(row=1, column=7).value = "Open"
wksh.cell(row=1, column=8).value = "Pre.Close"

i=0
for row in wksh.rows:
     i+=1
     if (i == 1): # skip the 1st row, it contains headings
        continue
        
     # Wait for the Quotes & Tools menu item to appear
     wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#nav_quotes_main > a:nth-child(1)")))
     driver.find_element_by_css_selector("#nav_quotes_main > a:nth-child(1)").click() # click on Quotes & Tools
     # Wait for the Symbol text box to appear
     wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.paddingLeft25:nth-child(2) > input:nth-child(2)")))
     symbol = driver.find_element_by_css_selector("div.paddingLeft25:nth-child(2) > input:nth-child(2)")
        
     test_symbol = wksh.cell(row =i, column =1).value
     symbol.send_keys(test_symbol, Keys.RETURN) # enter a symbol and get a quote
      
     # get a quote but wait for the "Summary" tab to appear
     wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".page_navigation > div:nth-child(1) > a:nth-child(1)")))
        
     last_price = driver.find_element_by_css_selector(".brownContainer > div:nth-child(1) > span:nth-child(1)").get_attribute("innerText")
     bid_price = driver.find_element_by_css_selector("#table1 > table > tbody > tr.tableRowBG > td:nth-child(2)").get_attribute("innerText")
     ask_price = driver.find_element_by_css_selector("#table1 > table > tbody > tr:nth-child(2) > td:nth-child(2)").get_attribute("innerText")
     high = driver.find_element_by_css_selector("#table1 > table > tbody > tr.tableRowBG > td:nth-child(4)").get_attribute("innerText")
     low =  driver.find_element_by_css_selector("#table1 > table > tbody > tr:nth-child(2) > td:nth-child(4)").get_attribute("innerText")
     open_price = driver.find_element_by_css_selector("#table1 > table > tbody > tr.tableRowBG > td:nth-child(6)").get_attribute("innerText")
     prev_close = driver.find_element_by_css_selector("#table1 > table > tbody > tr:nth-child(2) > td:nth-child(6)").get_attribute("innerText")
        
     wksh.cell(row = i, column =2).value = last_price
     wksh.cell(row = i, column =3).value =bid_price
     wksh.cell(row = i, column =4).value =ask_price
     wksh.cell(row = i, column =5).value =high
     wksh.cell(row = i, column =6).value =low
     wksh.cell(row = i, column =7).value =open_price
     wksh.cell(row = i, column =8).value =prev_close
    
wb.remove_sheet(wb.get_sheet_by_name('test_parameters')) # delete the test_parameters, no need to save them
wb.save(r"C:\temp\test2.xlsx")
        
print ("Test ending ...")
driver.quit()