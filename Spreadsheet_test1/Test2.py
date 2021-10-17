'''
Created on Jun04, 2020

@author: Home
'''
import openpyxl

wb_in = openpyxl.load_workbook('C:/temp/test-parameters.xlsx')
print ("Sheet names:", wb_in.sheetnames)

wksh_in = wb_in['symbols'] 
wb_out = openpyxl.Workbook()
wksh_out = wb_out.active
wksh_out.title = "Output test"

print ("Input sheet test1:")

for index, row in enumerate(wksh_in.rows):
    print ("Row: ", row, " ", row[0].value, row[1].value)
    print (wksh_in.cell(index+1, 1).value)
    print (wksh_in.cell(index+1, 2).value)
    wksh_out.cell(index+1,1).value = wksh_in.cell(index+1,1).value
    wksh_out.cell(index+1, 2).value = wksh_in.cell(index+1, 2).value
    wksh_out.cell(index+1, 3).value = (wksh_in.cell(index+1, 2)).value *2
    
print ("***\n")

print ("Output sheet test2, sheet by column:")
for col in wksh_out.columns:
    for cell in col:
        print (cell, cell.value)

wb_out.save("C:/temp/delete-me.xlsx")

print ("***\n")
print ("Output sheet test3, sheet by row:")
for row in wksh_in.rows:
    for cell in row:
        print (cell.value)