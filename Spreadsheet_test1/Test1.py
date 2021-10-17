'''
Created on Mar 27, 2020

@author: Home
'''
import openpyxl

wb = openpyxl.load_workbook('C:/temp/test-parameters.xlsx')
print ("Sheet names:", wb.sheetnames)

wksh = wb['test_parameters'] 

cell = wksh['A1'].value
print (cell)

print ("Input sheet:")
i=0  
for row in wksh.rows:
    i+=1
    print ("Row: ", row[0].value)
    symbol = wksh.cell(row =i, column =1).value
    price = (wksh.cell(row =i, column = 2).value)
    print (symbol,  price)
    wksh.cell(row = i, column =3).value = price *2
    
print ("\n***\n")

print ("Output sheet:")
for col in wksh.columns:
    for cell in col:
        print (cell, cell.value)

wb.save("C:/temp/delete-me.xlsx")